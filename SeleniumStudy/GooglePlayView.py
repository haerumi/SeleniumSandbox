# -*- coding: utf-8 -*-
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
from datetime import datetime
import os
import sys


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


template = resource_path("gpv_crawling_template.xlsx")
write_wb = load_workbook(template, data_only=True)
write_ws = write_wb['Sheet1']

popular_app_game_url = input(u"앱 순위를 가져올 url 을 입력하세요 : ")
ranking_count = input(u"가져올 앱 순위의 개수를 입력하세요 (기본값 : 100, 최대값 : 200) : ")

if ranking_count == '':
    ranking_count = 100
elif int(ranking_count) > 200:
    print(u"Play store 웹에서는 200위까지만 추출 가능합니다.")
    ranking_count = 200

options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-logging"])

if os.path.isfile("./chromedriver.exe"):

    driver = webdriver.Chrome("chromedriver.exe", chrome_options=options)
    driver.implicitly_wait(3)
    driver.get(popular_app_game_url)

    date = datetime.today().strftime(u"%Y-%m-%d %H:%M:%S")
    write_ws.cell(row=2, column=3, value=popular_app_game_url)
    write_ws.cell(row=3, column=3, value=date)

    checksum = False

    for i in range(1, int(ranking_count)+1):
        if not checksum:
            try:
                href_css = '#fcxH9b > div.WpDbMd > c-wiz > div > c-wiz > div > c-wiz > c-wiz > c-wiz > div > div.ZmHEEd > div:nth-child(' + str(
                    i) + ') > c-wiz > div > div > div.RZEgze > div > div > div.bQVA0c > div > div > div.b8cIId.ReQCgd.Q9MA7b > a'

                elem = driver.find_element_by_css_selector(href_css)
                elem2 = driver.find_element_by_css_selector(href_css + " > div")

            except NoSuchElementException:
                checksum = True
                href_css = '#fcxH9b > div.WpDbMd > c-wiz > div > c-wiz > div > c-wiz > c-wiz > c-wiz > div > div.ZmHEEd > c-wiz:nth-child(' + str(
                    i) + ') > div > div > div.RZEgze > div > div > div.bQVA0c > div > div > div.b8cIId.ReQCgd.Q9MA7b > a'
                try:
                    elem = driver.find_element_by_css_selector(href_css)
                    elem2 = driver.find_element_by_css_selector(href_css + " > div")

                except NoSuchElementException:
                    print(u"NoSuchElementException 발생, 재실행하거나 담당자에게 문의하세요")
                    break
        else:
            href_css = '#fcxH9b > div.WpDbMd > c-wiz > div > c-wiz > div > c-wiz > c-wiz > c-wiz > div > div.ZmHEEd > c-wiz:nth-child(' + str(
                i) + ') > div > div > div.RZEgze > div > div > div.bQVA0c > div > div > div.b8cIId.ReQCgd.Q9MA7b > a'

            elem = driver.find_element_by_css_selector(href_css)
            elem2 = driver.find_element_by_css_selector(href_css + " > div")

        href_str = str(elem.get_attribute("href"))
        href_str = href_str.split("id=")[1]

        write_ws.cell(row=i + 5, column=3, value=elem2.get_attribute("title"))
        write_ws.cell(row=i + 5, column=4, value=href_str)

        if i % 30 == 0:
            elem.send_keys(Keys.PAGE_DOWN)
            elem.send_keys(Keys.PAGE_DOWN)
            time.sleep(3)

    else:
        date = datetime.today().strftime("%Y%m%d_%H%M%S")
        write_wb.save("gpv_" + date + ".xlsx")
        driver.close()

        print(u"인기 앱 추출 완료!")
        print(u"실행 파일 위치에 저장된 excel 파일을 확인해주세요" + "(gpv_" + date + ".xlsx)")

else:
    print(u"실행 파일과 동일한 위치에 chromedriver.exe 를 저장해주세요")