# -*- coding: utf-8 -*-
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import pyautogui
from datetime import datetime
import os
import sys

write_wb = load_workbook("gpv_crawling_template.xlsx", data_only=True)
write_ws = write_wb['Sheet1']

popular_app_game_url = input(u"앱 순위를 가져올 url 을 입력하세요 : ")

driver = webdriver.Chrome("chromedriver.exe")
driver.implicitly_wait(3)

#popular_app_game_url = "https://play.google.com/store/apps/collection/cluster?clp=0g4jCiEKG3RvcHNlbGxpbmdfZnJlZV9BUFBMSUNBVElPThAHGAM%3D:S:ANO1ljKs-KA&gsr=CibSDiMKIQobdG9wc2VsbGluZ19mcmVlX0FQUExJQ0FUSU9OEAcYAw%3D%3D:S:ANO1ljL40zU"
driver.get(popular_app_game_url)

date = datetime.today().strftime("%Y%m%d_%H%M%S")
write_ws.cell(row=2, column=3, value=popular_app_game_url)
write_ws.cell(row=3, column=3, value=date)

for i in range(1, 101):

    try:
        href_css = '#fcxH9b > div.WpDbMd > c-wiz > div > c-wiz > div > c-wiz > c-wiz > c-wiz > div > div.ZmHEEd > div:nth-child(' + str(
            i) + ') > c-wiz > div > div > div.RZEgze > div > div > div.bQVA0c > div > div > div.b8cIId.ReQCgd.Q9MA7b > a'

        elem = driver.find_element_by_css_selector(href_css)
        elem2 = driver.find_element_by_css_selector(href_css + " > div")

    except Exception as e:
        href_css = '#fcxH9b > div.WpDbMd > c-wiz > div > c-wiz > div > c-wiz > c-wiz > c-wiz > div > div.ZmHEEd > c-wiz:nth-child(' + str(
            i) + ') > div > div > div.RZEgze > div > div > div.bQVA0c > div > div > div.b8cIId.ReQCgd.Q9MA7b > a'

        elem = driver.find_element_by_css_selector(href_css)
        elem2 = driver.find_element_by_css_selector(href_css + " > div")

    href_str = str(elem.get_attribute("href"))
    href_str = href_str.split("id=")[1]

    write_ws.cell(row=i+5, column=3, value=elem2.get_attribute("title"))
    write_ws.cell(row=i+5, column=4, value=href_str)

    if i % 20 == 0:
        elem.send_keys(Keys.PAGE_DOWN)
        elem.send_keys(Keys.PAGE_DOWN)
        time.sleep(3)

write_wb.save("C:\chromedriver_win32\st\stest_" + date + ".xlsx")
driver.close()
